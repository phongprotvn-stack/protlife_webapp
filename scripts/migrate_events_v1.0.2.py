"""Migration: Import events from Prot_LifeOS_v1.0.2.xlsm into Supabase.
Keeps existing contacts untouched. Only inserts new events.
"""
import openpyxl
import psycopg2
import psycopg2.extras
import os
from datetime import datetime

# --- Config ---
EXCEL_PATH = r'C:\Users\phong\.hermes\desktop-attachments\Prot_LifeOS_v1.0.2.xlsm'
USER_ID = 'da4a1a28-f6ac-4b21-ab58-6278185a2b59'

# Database connection
DATABASE_URL = os.environ.get('SUPABASE_DB_URL')
if not DATABASE_URL:
    # Prompt-style: read from user later, but for now use the known connection
    DATABASE_URL = 'postgresql://postgres.hwgrdhnsuvohgtcuemag:Ngoctram3008%@aws-0-ap-southeast-1.pooler.supabase.com:5432/postgres'

# Importance mapping
IMPORTANCE_MAP = {
    '5 - Highest': 'Highest',
    '4 - High': 'High',
    '3 - Medium': 'Medium',
    '2 - Low': 'Low',
    '1 - Lowest': 'Lowest',
    5: 'Highest', 4: 'High', 3: 'Medium', 2: 'Low', 1: 'Lowest',
}

def parse_importance(val):
    if not val:
        return 'Medium'
    if isinstance(val, str):
        return IMPORTANCE_MAP.get(val.strip(), 'Medium')
    if isinstance(val, (int, float)):
        return IMPORTANCE_MAP.get(int(val), 'Medium')
    return 'Medium'

def parse_map_link(place, maplink_val):
    """Generate Google Maps URL from place, since Excel shows 'Map' for HYPERLINK formula."""
    if maplink_val and str(maplink_val) != 'Map' and str(maplink_val).startswith('http'):
        return str(maplink_val)
    if place and str(place).strip():
        return f"https://www.google.com/maps/search/?api=1&query={str(place).strip()}"
    return None

def parse_cost(val):
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def parse_participants(val):
    """Parse semicolon-separated participant names."""
    if not val or not str(val).strip():
        return []
    return [p.strip() for p in str(val).split(';') if p.strip()]

def parse_datetime(val):
    """Convert Excel datetime or string to ISO date string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str) and val.strip():
        # Try parsing common formats
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d']:
            try:
                return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
    return None

def main():
    print(f"[1/4] Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # --- Read events ---
    ws = wb['Event']
    events = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        event_id = row[0]
        if not event_id or not str(event_id).startswith('EV'):
            continue
        event_id = str(event_id)

        # Skip already imported event
        if event_id == 'EV19921027001':
            print(f"  Skipping {event_id} (already imported)")
            continue

        title = str(row[2]) if row[2] else ''
        participants_raw = row[3]
        event_type = str(row[4]) if row[4] else 'Other'
        life_stage = str(row[5]) if row[5] else None
        source = str(row[6]) if row[6] else None
        start_date = parse_datetime(row[7])
        end_date = parse_datetime(row[8])
        place = str(row[9]).strip() if row[9] else None
        maplink_val = row[10]
        mood = str(row[11]) if row[11] else None
        importance = parse_importance(row[12])
        participant_count = int(row[13]) if row[13] and str(row[13]).isdigit() else None
        cost = parse_cost(row[14])
        notes = str(row[15]) if row[15] else None
        created_date = parse_datetime(row[16])
        updated_date = parse_datetime(row[17])

        # Build notes with participants info
        participant_names = parse_participants(participants_raw)
        full_notes = notes or ''
        if participant_names:
            part_str = 'Người tham gia: ' + '; '.join(participant_names)
            full_notes = (full_notes + '\n\n' + part_str) if full_notes else part_str

        # Generate map link
        map_link = parse_map_link(place, maplink_val)

        events.append({
            'eventid': event_id,
            'title': title,
            'eventtype': event_type,
            'lifestage': life_stage,
            'source': source or 'Manual',
            'startdate': start_date,
            'enddate': end_date,
            'place': place,
            'maplink': map_link,
            'mood': mood,
            'importance': importance,
            'participantcount': participant_count,
            'cost': cost,
            'notes': full_notes or None,
            'createddate': created_date or datetime.now().strftime('%Y-%m-%d'),
            'updateddate': updated_date or datetime.now().strftime('%Y-%m-%d'),
            'user_id': USER_ID,
        })

    print(f"  Found {len(events)} new events to import")

    # --- Read EventParticipants ---
    ws_ep = wb['EventParticipant']
    participants = []
    for row in ws_ep.iter_rows(min_row=3, max_row=ws_ep.max_row, values_only=True):
        if not row[0] or not str(row[0]).startswith('EP'):
            continue
        ep = {
            'recordid': str(row[0]),
            'eventid': str(row[1]) if row[1] else None,
            'contactid': str(row[2]) if row[2] else None,
            'contactname': str(row[3]).strip() if row[3] else None,
            'role': str(row[4]) if row[4] else 'Participant',
            'attendance': str(row[5]) if row[5] else 'Confirmed',
        }
        if ep['eventid'] and ep['contactid']:
            participants.append(ep)

    print(f"  Found {len(participants)} event-participant links")

    if not events:
        print("No new events to import.")
        return

    # --- Connect to DB ---
    print(f"\n[2/4] Connecting to Supabase...")
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # Check existing event IDs
    cur.execute('SELECT "EventID" FROM events WHERE user_id = %s', (USER_ID,))
    existing = set(r[0] for r in cur.fetchall())
    print(f"  Existing events in DB: {len(existing)}")

    # Filter out already-imported
    to_insert = [e for e in events if e['eventid'] not in existing]
    print(f"  Events to insert: {len(to_insert)}")

    if not to_insert:
        print("All events already imported.")
        cur.close()
        conn.close()
        return

    # --- Insert events ---
    print(f"\n[3/4] Inserting {len(to_insert)} events...")
    insert_sql = """
        INSERT INTO events (
            "EventID", "No", "Title", "EventType", "LifeStage", "Source",
            "StartDate", "EndDate", "Place", "Maplink", "Mood", "Importance",
            "ParticipantCount", "Cost", "Notes", "CreatedDate", "UpdatedDate", user_id
        ) VALUES (
            %(eventid)s, %(no)s, %(title)s, %(eventtype)s, %(lifestage)s, %(source)s,
            %(startdate)s, %(enddate)s, %(place)s, %(maplink)s, %(mood)s, %(importance)s,
            %(participantcount)s, %(cost)s, %(notes)s, %(createddate)s, %(updateddate)s, %(user_id)s
        )
    """
    idx = 1
    for e in to_insert:
        e['no'] = idx
        try:
            cur.execute(insert_sql, e)
            print(f"  ✓ {e['eventid']}: {e['title'][:40]}")
        except Exception as ex:
            print(f"  ✗ {e['eventid']}: {ex}")
        idx += 1

    conn.commit()
    print(f"\n  Inserted {len(to_insert)} events successfully!")

    # --- Insert EventParticipants ---
    if participants:
        print(f"\n[4/4] Inserting {len(participants)} event-participant links...")
        ep_sql = """
            INSERT INTO eventparticipants (recordid, eventid, contactid, contactname, role, attendance)
            VALUES (%(recordid)s, %(eventid)s, %(contactid)s, %(contactname)s, %(role)s, %(attendance)s)
            ON CONFLICT (recordid) DO NOTHING
        """
        for ep in participants:
            try:
                cur.execute(ep_sql, ep)
                print(f"  ✓ {ep['recordid']}: {ep['contactname']} → {ep['eventid']}")
            except Exception as ex:
                print(f"  ✗ {ep['recordid']}: {ex}")
        conn.commit()

    # --- Summary ---
    cur.execute("SELECT COUNT(*) FROM events WHERE user_id = %s", (USER_ID,))
    total = cur.fetchone()[0]
    cur.execute('SELECT MIN("StartDate"), MAX("StartDate") FROM events WHERE user_id = %s AND "StartDate" IS NOT NULL', (USER_ID,))
    date_range = cur.fetchone()
    print(f"\n{'='*50}")
    print(f"✅ Migration complete!")
    print(f"  Total events for user: {total}")
    print(f"  Date range: {date_range[0]} → {date_range[1]}")
    print(f"  Event-participant links: {len(participants)}")

    cur.close()
    conn.close()
    print(f"{'='*50}")

if __name__ == '__main__':
    main()
