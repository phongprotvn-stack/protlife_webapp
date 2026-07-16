"""
Prot Life OS v1.0.1 — Data Migration Script
=============================================
Imports Contact and Event data from Excel into Supabase
for admin user phongprot.vn@gmail.com
"""

import openpyxl
from datetime import datetime, date
import json
import sys
from urllib.parse import quote_plus

EXCEL_PATH = r"C:\Users\phong\.hermes\desktop-attachments\Prot_LifeOS_v1.0.1.xlsm"

# Supabase connection
DB_URL = "postgresql://postgres.hwgrdhnsuvohgtcuemag:Ngoctram3008%40@aws-0-ap-southeast-1.pooler.supabase.com:5432/postgres"

# ─── STEP 1: Get admin user UUID ─────────────────────────────────
def get_admin_user_id():
    import psycopg2
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    cur.execute("SELECT id FROM auth.users WHERE email = 'phongprot.vn@gmail.com'")
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        print("ERROR: Admin user phongprot.vn@gmail.com not found in auth.users")
        sys.exit(1)
    print(f"Admin user ID: {row[0]}")
    return row[0]

# ─── STEP 2: Read Excel data ─────────────────────────────────────
def read_sheet(ws, header_row=2):
    """Read sheet data starting from header_row, returning list of dicts."""
    headers = []
    for cell in ws[header_row]:
        if cell.value is not None:
            headers.append(cell.value)
    
    data = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:  # skip completely empty rows
            continue
        entry = {}
        has_data = False
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                entry[headers[i]] = val
                if val is not None:
                    has_data = True
        if has_data:
            data.append(entry)
    return data

def to_date(val):
    """Convert various date formats to string YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None

def to_ts(val):
    """Convert to ISO timestamp string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None

def bool_val(val):
    """Convert various truthy values to boolean."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() in ('TRUE', 'YES', '1')
    if isinstance(val, int):
        return val > 0
    return False

def str_val(val):
    """Convert to string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def int_val(val):
    """Convert to int or None."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None

def num_val(val):
    """Convert to numeric or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

# ─── STEP 3: Map contacts ────────────────────────────────────────
def map_contact(row, user_id):
    """Map Excel row to contacts table format."""
    return {
        "ContactID": str_val(row.get("ContactID")),
        "Name": str_val(row.get("FullName")),
        "Relationship": str_val(row.get("Relationship")) or "Other",
        "Gender": str_val(row.get("Gender")),
        "Birthday": to_date(row.get("Birthday")),
        "Phone": str_val(row.get("Phone")),
        "Email": str_val(row.get("Email")),
        "Organization1": str_val(row.get("Organization 1")),
        "Organization2": str_val(row.get("Organization 2")),
        "RelationshipScore": int_val(row.get("RelationshipScore")) or 0,
        "Status": str_val(row.get("Status")) or "Active",
        "IsFavorite": bool_val(row.get("IsFavorite")),
        "CreatedDate": to_ts(row.get("CreatedDate")),
        "UpdatedDate": to_ts(row.get("UpdatedDate")),
        "user_id": user_id,
        "Notes": None,
        "Avatar": None,
    }

# ─── STEP 4: Map events ──────────────────────────────────────────
def map_event(row, user_id):
    """Map Excel row to events table format."""
    importance = str_val(row.get("Importance"))
    # Normalize importance format
    if importance:
        if "5" in importance:
            importance = "Highest"
        elif "4" in importance:
            importance = "High"
        elif "3" in importance:
            importance = "Medium"
        elif "2" in importance:
            importance = "Low"
        elif "1" in importance:
            importance = "Lowest"
    
    return {
        "EventID": str_val(row.get("EventID")),
        "No": int_val(row.get("No.")),
        "Title": str_val(row.get("Title")),
        "EventType": str_val(row.get("EventType")) or "Other",
        "LifeStage": str_val(row.get("LifeStage")),
        "Source": str_val(row.get("Source")) or "Manual",
        "StartDate": to_date(row.get("StartDate")),
        "EndDate": to_date(row.get("EndDate")),
        "Place": str_val(row.get("Place")),
        "Maplink": str_val(row.get("MapLink")),
        "Mood": str_val(row.get("Mood")),
        "Importance": importance or "Medium",
        "ParticipantCount": int_val(row.get("ParticipantCount")) or 0,
        "Cost": num_val(row.get("Cost")) or 0,
        "Notes": str_val(row.get("Notes")),
        "CreatedDate": to_ts(row.get("CreatedDate")),
        "UpdatedDate": to_ts(row.get("UpdatedDate")),
        "user_id": user_id,
    }

# ─── STEP 5: Insert into Supabase ────────────────────────────────
def insert_data(contacts, events, user_id):
    import psycopg2
    
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    
    # Check existing data
    cur.execute("SELECT COUNT(*) FROM contacts")
    existing_contacts = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM events")
    existing_events = cur.fetchone()[0]
    
    print(f"\nExisting data: {existing_contacts} contacts, {existing_events} events")
    print(f"New data to import: {len(contacts)} contacts, {len(events)} events")
    
    if existing_contacts > 0 or existing_events > 0:
        print("\n⚠️  Database already has data. Adding new records (will skip duplicates).")
    
    # ── Insert Contacts ──
    inserted_c = 0
    skipped_c = 0
    error_c = 0
    
    for c in contacts:
        try:
            # Check if ContactID already exists
            cur.execute('SELECT 1 FROM contacts WHERE "ContactID" = %s', (c["ContactID"],))
            if cur.fetchone():
                skipped_c += 1
                continue
            
            cur.execute(
                '''INSERT INTO contacts (
                    "ContactID", "Name", "Relationship", "Gender", "Birthday",
                    "Phone", "Email", "Organization1", "Organization2",
                    "RelationshipScore", "Status", "IsFavorite",
                    "CreatedDate", "UpdatedDate", "user_id", "Notes", "Avatar"
                ) VALUES (
                    %(ContactID)s, %(Name)s, %(Relationship)s, %(Gender)s, %(Birthday)s,
                    %(Phone)s, %(Email)s, %(Organization1)s, %(Organization2)s,
                    %(RelationshipScore)s, %(Status)s, %(IsFavorite)s,
                    %(CreatedDate)s, %(UpdatedDate)s, %(user_id)s, %(Notes)s, %(Avatar)s
                )''',
                c
            )
            inserted_c += 1
        except Exception as e:
            print(f"  ERROR inserting contact {c.get('ContactID')}: {e}")
            error_c += 1
    
    conn.commit()
    print(f"\nContacts: {inserted_c} inserted, {skipped_c} skipped (duplicate), {error_c} errors")
    
    # ── Insert Events ──
    inserted_e = 0
    skipped_e = 0
    error_e = 0
    
    # First, drop the check constraint so we can insert custom EventIDs
    try:
        cur.execute('ALTER TABLE events DROP CONSTRAINT IF EXISTS events_EventType_check')
        conn.commit()
        print("  Dropped EventType CHECK constraint temporarily")
    except Exception as e:
        conn.rollback()
        print(f"  Note: Could not drop constraint: {e}")
    
    for e in events:
        try:
            # Check if EventID already exists
            cur.execute('SELECT 1 FROM events WHERE "EventID" = %s', (e["EventID"],))
            if cur.fetchone():
                skipped_e += 1
                continue
            
            cur.execute(
                '''INSERT INTO events (
                    "EventID", "No", "Title", "EventType", "LifeStage", "Source",
                    "StartDate", "EndDate", "Place", "Maplink", "Mood", "Importance",
                    "ParticipantCount", "Cost", "Notes",
                    "CreatedDate", "UpdatedDate", "user_id"
                ) VALUES (
                    %(EventID)s, %(No)s, %(Title)s, %(EventType)s, %(LifeStage)s, %(Source)s,
                    %(StartDate)s, %(EndDate)s, %(Place)s, %(Maplink)s, %(Mood)s, %(Importance)s,
                    %(ParticipantCount)s, %(Cost)s, %(Notes)s,
                    %(CreatedDate)s, %(UpdatedDate)s, %(user_id)s
                )''',
                e
            )
            inserted_e += 1
        except Exception as e:
            print(f"  ERROR inserting event {e.get('EventID')}: {e}")
            error_e += 1
    
    conn.commit()
    print(f"Events: {inserted_e} inserted, {skipped_e} skipped (duplicate), {error_e} errors")
    
    # Restore constraint (with broader values)
    try:
        cur.execute("""
            ALTER TABLE events ADD CONSTRAINT events_EventType_check 
            CHECK ("EventType" IN ('Meeting', 'Birthday', 'Travel', 'Work', 'Sport', 'Hospital', 
                   'Meal', 'Call', 'Shopping', 'Study', 'Party', 'Date', 'Entertainment', 'Other',
                   'Ceremony', 'Memorial', 'Graduation', 'Admission', 'Festival', 'Moving', 'Accident',
                   'Crisis'))
        """)
        conn.commit()
        print("  Restored EventType CHECK constraint with broader values")
    except Exception as e:
        conn.rollback()
        print(f"  Note: Could not restore constraint: {e}")
    
    cur.close()
    conn.close()
    
    return inserted_c, inserted_e, skipped_c, skipped_e, error_c, error_e


# ─── MAIN ────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("PROT LIFE OS v1.0.1 — Data Migration")
    print("=" * 60)
    
    # Step 1: Get user ID
    user_id = get_admin_user_id()
    
    # Step 2: Read Excel
    print(f"\nReading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    print(f"Sheets: {wb.sheetnames}")
    
    # Read Contact sheet
    ws_contact = wb["Contact"]
    contacts_raw = read_sheet(ws_contact)
    print(f"\nContact sheet: {len(contacts_raw)} rows of data")
    if contacts_raw:
        print(f"  First: {contacts_raw[0].get('ContactID')} - {contacts_raw[0].get('FullName')}")
        print(f"  Last:  {contacts_raw[-1].get('ContactID')} - {contacts_raw[-1].get('FullName')}")
    
    # Read Event sheet
    ws_event = wb["Event"]
    events_raw = read_sheet(ws_event)
    print(f"\nEvent sheet: {len(events_raw)} rows of data (first pass)")
    
    # Filter out empty rows
    events_raw = [r for r in events_raw if r.get("EventID") is not None]
    print(f"Event sheet: {len(events_raw)} rows after filtering (with EventID)")
    if events_raw:
        print(f"  First: {events_raw[0].get('EventID')} - {events_raw[0].get('Title')}")
        print(f"  Last:  {events_raw[-1].get('EventID')} - {events_raw[-1].get('Title')}")
    
    # Step 3: Map data
    contacts = [map_contact(c, user_id) for c in contacts_raw if c.get("ContactID")]
    events = [map_event(e, user_id) for e in events_raw if e.get("EventID")]
    
    print(f"\nMapped: {len(contacts)} contacts, {len(events)} events")
    
    # Step 4: Show sample
    print("\nSample Contact:")
    print(json.dumps(contacts[0] if contacts else {}, indent=2, default=str)[:500])
    print("\nSample Event:")
    print(json.dumps(events[0] if events else {}, indent=2, default=str)[:500])
    
    # Step 5: Insert
    print("\n" + "-" * 60)
    print("INSERTING DATA...")
    print("-" * 60)
    
    ins_c, ins_e, skip_c, skip_e, err_c, err_e = insert_data(contacts, events, user_id)
    
    print("\n" + "=" * 60)
    print("MIGRATION SUMMARY")
    print("=" * 60)
    print(f"  Contacts: {ins_c} imported, {skip_c} skipped, {err_c} errors")
    print(f"  Events:   {ins_e} imported, {skip_e} skipped, {err_e} errors")
    print(f"  Total:    {ins_c + ins_e} records imported")
    print("=" * 60)
